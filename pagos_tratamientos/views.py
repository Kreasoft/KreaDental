from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from django.db import models
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import PagoTratamiento, HistorialPago
from tratamientos.models import Tratamiento
from formas_pago.models import FormaPago
from django.contrib.humanize.templatetags.humanize import intcomma

@login_required
def lista_pagos(request):
    pagos = PagoTratamiento.objects.all().order_by('-fecha_pago')
    
    # Filtrar por tratamiento si se proporciona el ID
    tratamiento_id = request.GET.get('tratamiento')
    if tratamiento_id:
        pagos = pagos.filter(tratamiento_id=tratamiento_id)
        tratamiento = get_object_or_404(Tratamiento, id=tratamiento_id)
        tratamientos_disponibles = None  # No mostrar la lista de tratamientos si ya se está viendo uno específico
    else:
        tratamiento = None
        # Obtener tratamientos que tengan saldo pendiente (más flexible)
        tratamientos_disponibles = Tratamiento.objects.annotate(
            total_pagado=Sum('pagos__monto', filter=models.Q(pagos__estado='COMPLETADO'))
        ).annotate(
            saldo_pendiente=models.F('costo_total') - Coalesce(models.F('total_pagado'), 0, output_field=models.DecimalField())
        ).filter(
            saldo_pendiente__gt=0
        ).exclude(
            estado='COMPLETADO'  # Excluir tratamientos completados
        ).order_by('-fecha_creacion')
    
    # Estadísticas
    total_pagos = pagos.filter(estado='COMPLETADO').aggregate(total=Sum('monto'))['total'] or 0
    pagos_pendientes = pagos.filter(estado='PENDIENTE').count()
    pagos_completados = pagos.filter(estado='COMPLETADO').count()
    pagos_anulados = pagos.filter(estado='ANULADO').count()
    
    # Formatear números para el contexto
    total_pagos_formatted = intcomma(int(total_pagos)) if total_pagos else '0'
    
    context = {
        'pagos': pagos,
        'total_pagos': total_pagos,
        'total_pagos_formatted': total_pagos_formatted,
        'pagos_pendientes': pagos_pendientes,
        'pagos_completados': pagos_completados,
        'pagos_anulados': pagos_anulados,
        'tratamiento': tratamiento,
        'tratamientos_disponibles': tratamientos_disponibles
    }
    
    return render(request, 'pagos_tratamientos/lista_pagos.html', context)

@login_required
def crear_pago(request, tratamiento_id):
    tratamiento = get_object_or_404(Tratamiento, id=tratamiento_id)
    formas_pago = FormaPago.objects.filter(estado=True)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '').strip()
            
            # Limpiar el valor del monto (remover espacios y convertir coma a punto)
            monto_str = monto_str.replace(' ', '').replace(',', '.')
            
            # Validar que sea un número válido
            if not monto_str or not monto_str.replace('.', '').replace('-', '').isdigit():
                raise ValueError("El monto debe ser un número válido")
            
            monto = float(monto_str)
            
            if monto < 0:
                raise ValueError("El monto no puede ser negativo")
            
            forma_pago_id = request.POST.get('forma_pago')
            comprobante = request.POST.get('comprobante', '')
            notas = request.POST.get('notas', '')
            
            forma_pago = get_object_or_404(FormaPago, id=forma_pago_id)
            
            pago = PagoTratamiento.objects.create(
                tratamiento=tratamiento,
                forma_pago=forma_pago,
                monto=monto,
                comprobante=comprobante,
                notas=notas,
                creado_por=request.user,
                estado='COMPLETADO'
            )
            
            # Crear registro en el historial
            HistorialPago.objects.create(
                pago=pago,
                tipo_accion='CREACION',
                estado_nuevo='COMPLETADO',
                monto_nuevo=monto,
                realizado_por=request.user
            )
            
            messages.success(request, 'Pago registrado exitosamente.')
            return redirect('tratamientos:detalle', pk=tratamiento.id)
            
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, 'Error al procesar el pago.')
    
    # Calcular saldo pendiente
    pagos_completados = tratamiento.pagos.filter(estado='COMPLETADO')
    total_pagado = pagos_completados.aggregate(total=Sum('monto'))['total'] or 0
    saldo_pendiente = tratamiento.costo_total - total_pagado
    
    context = {
        'tratamiento': tratamiento,
        'formas_pago': formas_pago,
        'saldo_pendiente': saldo_pendiente
    }
    
    return render(request, 'pagos_tratamientos/crear_pago.html', context)

@login_required
def anular_pago(request, pago_id):
    if request.method == 'POST':
        pago = get_object_or_404(PagoTratamiento, id=pago_id)
        
        if pago.estado == 'ANULADO':
            return JsonResponse({
                'status': 'error',
                'message': 'Este pago ya está anulado.'
            })
        
        try:
            estado_anterior = pago.estado
            monto_anterior = pago.monto
            
            pago.estado = 'ANULADO'
            pago.save()
            
            # Crear registro en el historial
            HistorialPago.objects.create(
                pago=pago,
                tipo_accion='ANULACION',
                estado_anterior=estado_anterior,
                estado_nuevo='ANULADO',
                monto_anterior=monto_anterior,
                monto_nuevo=pago.monto,
                realizado_por=request.user,
                notas=request.POST.get('motivo_anulacion', '')
            )
            
            return JsonResponse({
                'status': 'success',
                'message': 'Pago anulado exitosamente.'
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error al anular el pago: {str(e)}'
            })
    
    return JsonResponse({
        'status': 'error',
        'message': 'Método no permitido.'
    })

@login_required
def detalle_pago(request, pago_id):
    pago = get_object_or_404(PagoTratamiento, id=pago_id)
    historial = pago.historial.all().order_by('-fecha_accion')
    
    context = {
        'pago': pago,
        'historial': historial
    }
    
    return render(request, 'pagos_tratamientos/detalle_pago.html', context)

@login_required
def editar_pago(request, pago_id):
    pago = get_object_or_404(PagoTratamiento, id=pago_id)
    formas_pago = FormaPago.objects.filter(estado=True)
    
    # Calcular montos del tratamiento
    tratamiento = pago.tratamiento
    pagos_completados = tratamiento.pagos.filter(estado='COMPLETADO').exclude(id=pago.id)
    total_pagado = pagos_completados.aggregate(total=models.Sum('monto'))['total'] or 0
    saldo_pendiente = tratamiento.costo_total - total_pagado
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '').strip()
            
            # Limpiar el valor del monto (remover espacios y convertir coma a punto)
            monto_str = monto_str.replace(' ', '').replace(',', '.')
            
            # Validar que sea un número válido
            if not monto_str or not monto_str.replace('.', '').replace('-', '').isdigit():
                raise ValueError("El monto debe ser un número válido")
            
            monto = float(monto_str)
            
            if monto < 0:
                raise ValueError("El monto no puede ser negativo")
            
            forma_pago_id = request.POST.get('forma_pago')
            comprobante = request.POST.get('comprobante', '')
            notas = request.POST.get('notas', '')
            
            forma_pago = get_object_or_404(FormaPago, id=forma_pago_id)
            
            # Guardar valores anteriores para el historial
            estado_anterior = pago.estado
            monto_anterior = pago.monto
            forma_pago_anterior = pago.forma_pago
            
            # Actualizar el pago
            pago.forma_pago = forma_pago
            pago.monto = monto
            pago.comprobante = comprobante
            pago.notas = notas
            pago.save()
            
            # Crear registro en el historial
            HistorialPago.objects.create(
                pago=pago,
                tipo_accion='MODIFICACION',
                estado_anterior=estado_anterior,
                estado_nuevo=pago.estado,
                monto_anterior=monto_anterior,
                monto_nuevo=monto,
                realizado_por=request.user,
                notas=f'Forma de pago anterior: {forma_pago_anterior}, Nueva forma de pago: {forma_pago}'
            )
            
            messages.success(request, 'Pago actualizado exitosamente.')
            return redirect('pagos_tratamientos:detalle_pago', pago_id=pago.id)
            
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, 'Error al actualizar el pago.')
    
    context = {
        'pago': pago,
        'formas_pago': formas_pago,
        'tratamiento': tratamiento,
        'total_pagado': total_pagado,
        'saldo_pendiente': saldo_pendiente
    }
    
    return render(request, 'pagos_tratamientos/editar_pago.html', context)

@login_required
def exportar_pagos(request):
    # Obtener los pagos
    pagos = PagoTratamiento.objects.all().order_by('-fecha_pago')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"
    
    # Estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    
    # Encabezados
    headers = [
        'ID', 'Fecha', 'Paciente', 'Tratamiento', 'Monto', 
        'Método de Pago', 'Estado', 'Comprobante', 'Notas'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos
    for row_num, pago in enumerate(pagos, 2):
        ws.cell(row=row_num, column=1, value=pago.id)
        ws.cell(row=row_num, column=2, value=pago.fecha_pago.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=3, value=str(pago.tratamiento.paciente) if pago.tratamiento.paciente else '')
        ws.cell(row=row_num, column=4, value=pago.tratamiento.get_estado_display())
        ws.cell(row=row_num, column=5, value=float(pago.monto))
        ws.cell(row=row_num, column=6, value=str(pago.forma_pago) if pago.forma_pago else '')
        ws.cell(row=row_num, column=7, value=pago.get_estado_display())
        ws.cell(row=row_num, column=8, value=pago.comprobante or '')
        ws.cell(row=row_num, column=9, value=pago.notas or '')
    
    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pagos.xlsx'
    
    # Guardar el libro de trabajo en la respuesta
    wb.save(response)
    return response

@login_required
def resumen_tratamiento(request, pago_id):
    """Vista para obtener el resumen de un tratamiento y sus pagos"""
    pago = get_object_or_404(PagoTratamiento, id=pago_id)
    tratamiento = pago.tratamiento
    
    # Obtener todos los pagos del tratamiento
    pagos_tratamiento = tratamiento.pagos.filter(estado='COMPLETADO').order_by('fecha_pago')
    total_pagado = pagos_tratamiento.aggregate(total=Sum('monto'))['total'] or 0
    saldo_pendiente = tratamiento.costo_total - total_pagado
    
    context = {
        'tratamiento': tratamiento,
        'pagos': pagos_tratamiento,
        'total_pagado': total_pagado,
        'saldo_pendiente': saldo_pendiente,
        'pago_actual': pago
    }
    
    return render(request, 'pagos_tratamientos/resumen_tratamiento.html', context)
