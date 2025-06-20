from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from .models import Cita
from .forms import CitaForm
from profesionales.models import Profesional

@login_required
def lista_citas(request):
    citas = Cita.objects.select_related('paciente', 'profesional').all()
    profesionales = Profesional.objects.all()
    return render(request, 'citas/lista_citas.html', {
        'citas': citas,
        'profesionales': profesionales
    })

@login_required
def crear_cita(request):
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cita creada correctamente.')
            return redirect('citas:lista_citas')
    else:
        initial = {}
        if request.GET.get('fecha'):
            initial['fecha'] = request.GET.get('fecha')
        if request.GET.get('hora'):
            initial['hora'] = request.GET.get('hora')
        if request.GET.get('profesional'):
            initial['profesional'] = request.GET.get('profesional')
        form = CitaForm(initial=initial)
    
    return render(request, 'citas/form_cita.html', {
        'form': form,
        'accion': 'Crear'
    })

@login_required
def editar_cita(request, pk):
    cita = get_object_or_404(Cita, pk=pk)
    if request.method == 'POST':
        form = CitaForm(request.POST, instance=cita)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cita actualizada correctamente.')
            return redirect('citas:lista_citas')
    else:
        form = CitaForm(instance=cita)
    
    return render(request, 'citas/form_cita.html', {
        'form': form,
        'cita': cita,
        'accion': 'Editar'
    })

@login_required
def eliminar_cita(request, pk):
    cita = get_object_or_404(Cita, pk=pk)
    if request.method == 'POST':
        cita.delete()
        messages.success(request, 'Cita eliminada correctamente.')
        return redirect('citas:lista_citas')
    
    return render(request, 'citas/confirmar_eliminar.html', {
        'cita': cita
    })

@login_required
def informes(request):
    # Obtener fechas del filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    profesional_id = request.GET.get('profesional')
    
    # Consulta base
    citas = Cita.objects.select_related('paciente', 'profesional')
    
    # Aplicar filtros si existen
    if fecha_inicio:
        citas = citas.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        citas = citas.filter(fecha__lte=fecha_fin)
    if profesional_id:
        citas = citas.filter(profesional_id=profesional_id)
    
    # Estadísticas
    total_citas = citas.count()
    citas_por_estado = citas.values('estado').annotate(total=Count('id'))
    citas_por_profesional = citas.values('profesional__nombre').annotate(total=Count('id'))
    
    profesionales = Profesional.objects.all()
    
    context = {
        'citas': citas,
        'profesionales': profesionales,
        'total_citas': total_citas,
        'citas_por_estado': citas_por_estado,
        'citas_por_profesional': citas_por_profesional,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'profesional_id': profesional_id
    }
    
    # Si es una solicitud de exportación
    if request.GET.get('export') == 'excel':
        return exportar_excel(request, citas)
    
    return render(request, 'citas/informes.html', context)

def exportar_excel(request, citas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de Citas"
    
    # Encabezados
    headers = ['Fecha', 'Hora', 'Paciente', 'Profesional', 'Estado', 'Motivo']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos
    for row, cita in enumerate(citas, 2):
        ws.cell(row=row, column=1, value=cita.fecha.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=cita.hora.strftime('%H:%M'))
        ws.cell(row=row, column=3, value=str(cita.paciente))
        ws.cell(row=row, column=4, value=str(cita.profesional))
        ws.cell(row=row, column=5, value=cita.get_estado_display())
        ws.cell(row=row, column=6, value=cita.motivo)
    
    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=informe_citas.xlsx'
    wb.save(response)
    return response 